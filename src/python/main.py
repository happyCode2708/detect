import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PilImage
from io import BytesIO

def scan_images_and_create_sheet(directory, output_file):
    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    
    # Set the column titles
    sheet.append(["Folder Name", "File Name", "Image"])
    
    row_height = 100
    sheet.row_dimensions[1].height = row_height  # Set the height of the header row
    
    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff')):
                folder_name = os.path.basename(root)
                file_name = os.path.splitext(file)[0]
                file_path = os.path.join(root, file)
                
                # Append the folder and file name
                row = [folder_name, file_name]
                
                # Open and resize the image
                img = PilImage.open(file_path)
                img.thumbnail((500, 500))
                
                # Save the thumbnail to a BytesIO stream
                img_stream = BytesIO()
                img.save(img_stream, format='PNG')
                img_stream.seek(0)
                
                # Create an image object for openpyxl
                img_openpyxl = OpenpyxlImage(img_stream)
                img_openpyxl.width = img.width
                img_openpyxl.height = img.height
                
                # Append the row data to the sheet
                row_index = sheet.max_row + 1
                sheet.append(row)
                
                # Set the row height
                sheet.row_dimensions[row_index].height = row_height
                
                # Add the image to the sheet
                img_openpyxl.anchor = f'C{row_index}'
                sheet.add_image(img_openpyxl)
    
    # Save the workbook
    workbook.save(output_file)
    print(f"Workbook saved as {output_file}")

# Directory to scan
directory = r"C:\Git\Gvi\RIVIR_FE\Portal\Frontend\src\assets\Certifiers"
# Output file path
output_file = r"output.xlsx"

# Run the function
scan_images_and_create_sheet(directory, output_file)
